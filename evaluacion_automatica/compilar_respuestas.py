"""Compila las planillas de los jueces en artefactos publicables.

Uso: python evaluacion_automatica/compilar_respuestas.py planilla1.xlsx planilla2.xlsx ...
Genera evaluacion/respuestas-jueces.json (escribe en evaluacion/) (datos) y respuestas-jueces.md (lectura).

Cada juez completa la primera hoja de su copia de la planilla; se toma la hoja
con más celdas cargadas. Los nombres no se publican: los jueces quedan
numerados por orden de archivo. La columna "detalle" conserva la semántica del
encabezado de la planilla: qué es incorrecto en los ítems fijos, o el texto de
la consulta en las filas P1-P3.
"""
import json
import sys
from pathlib import Path

import openpyxl


def hoja_con_datos(wb):
    def llenas(ws):
        return sum(1 for fila in ws.iter_rows(min_row=2)
                   for c in fila if c.value is not None and c.column > 2)
    hojas = [wb[h] for h in wb.sheetnames if h.startswith('Juez')]
    return max(hojas, key=llenas)


def fila_a_item(ws, f):
    return {
        'utilidad': int(float(ws.cell(f, 2).value)),
        'fidelidad': int(float(ws.cell(f, 3).value)),
        'completitud': int(float(ws.cell(f, 4).value)),
        'claridad': int(float(ws.cell(f, 5).value)),
        'incorrecto': str(ws.cell(f, 6).value or '').strip().upper().startswith(('SI', 'SÍ')),
        'detalle': str(ws.cell(f, 7).value or '').strip() or None,
    }


def copia_anonima(ruta, n, destino):
    wb = openpyxl.load_workbook(ruta)
    for h in wb.sheetnames:
        if h.startswith('Juez') and str(wb[h].cell(1, 2).value or '').strip():
            wb[h].cell(1, 2).value = f'Juez {n}'
    # Los metadatos del archivo también identifican (autor de Excel).
    wb.properties.creator = None
    wb.properties.lastModifiedBy = None
    wb.save(destino)


def compilar(rutas):
    jueces = []
    for n, ruta in enumerate(sorted(rutas), start=1):
        ws = hoja_con_datos(openpyxl.load_workbook(ruta, data_only=True))
        j = {'juez': n, 'items': [], 'consultas_propias': [], 'comentario': None}
        for f in range(4, 19):
            j['items'].append({'item': int(ws.cell(f, 1).value), **fila_a_item(ws, f)})
        for f in range(19, 22):
            j['consultas_propias'].append({'id': str(ws.cell(f, 1).value), **fila_a_item(ws, f)})
        for f in range(22, ws.max_row + 1):
            if str(ws.cell(f, 1).value or '').startswith('Comentario'):
                j['comentario'] = str(ws.cell(f, 2).value or '').strip() or None
        jueces.append(j)
    return jueces


DIMS = ['utilidad', 'fidelidad', 'completitud', 'claridad']


def a_markdown(jueces):
    import statistics as st
    L = ['# Respuestas de la evaluación con jueces', '',
         'Instrumento: [cuestionario-humano.md](cuestionario-humano.md) · '
         'Ítems evaluados: [items-congelados.json](items-congelados.json) · '
         'Datos crudos: [respuestas-jueces.json](respuestas-jueces.json) · '
         'Planillas originales (anonimizadas): '
         + ', '.join(f'[juez {j["juez"]}](planilla-juez-{j["juez"]}.xlsx)' for j in jueces),
         '',
         f'## Resumen — ítems fijos ({len(jueces)} jueces, '
         f'n={15 * len(jueces)} por dimensión)', '',
         '| Dimensión | Media ± DE |', '|---|---|']
    for d in DIMS:
        v = [i[d] for j in jueces for i in j['items']]
        L.append(f'| {d.capitalize()} | {st.mean(v):.2f} ± {st.stdev(v):.2f} |')
    for j in jueces:
        L += ['', f'## Juez {j["juez"]}', '',
              '| Ítem | Utilidad | Fidelidad | Completitud | Claridad | ¿Incorrecto? | Detalle |',
              '|---|---|---|---|---|---|---|']
        for i in j['items']:
            L.append(f'| {i["item"]} | {i["utilidad"]} | {i["fidelidad"]} | '
                     f'{i["completitud"]} | {i["claridad"]} | '
                     f'{"SÍ" if i["incorrecto"] else "no"} | {i["detalle"] or ""} |')
        L += ['', 'Consultas propias (la columna Detalle registra la consulta '
              'realizada o la observación del juez):', '',
              '| | Utilidad | Fidelidad | Completitud | Claridad | ¿Incorrecto? | Detalle |',
              '|---|---|---|---|---|---|---|']
        for i in j['consultas_propias']:
            L.append(f'| {i["id"]} | {i["utilidad"]} | {i["fidelidad"]} | '
                     f'{i["completitud"]} | {i["claridad"]} | '
                     f'{"SÍ" if i["incorrecto"] else "no"} | {i["detalle"] or ""} |')
        if j['comentario']:
            L += ['', f'Comentario libre: *{j["comentario"]}*']
    return '\n'.join(L) + '\n'


if __name__ == '__main__':
    carpeta = Path(__file__).parent.parent / 'evaluacion'
    datos = compilar(sys.argv[1:])
    for n, ruta in enumerate(sorted(sys.argv[1:]), start=1):
        copia_anonima(ruta, n, carpeta / f'planilla-juez-{n}.xlsx')
    (carpeta / 'respuestas-jueces.json').write_text(
        json.dumps(datos, ensure_ascii=False, indent=1), encoding='utf-8')
    (carpeta / 'respuestas-jueces.md').write_text(a_markdown(datos), encoding='utf-8')
    print(f'{len(datos)} jueces -> respuestas-jueces.json + respuestas-jueces.md')
